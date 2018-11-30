from openpyxl import Workbook
from openpyxl.styles import Font, colors
from openpyxl.chart import PieChart, Reference, BarChart
from openpyxl.chart.series import DataPoint
from collections import OrderedDict


class ReportCreator():

    def __init__(self, report):
        self.report = OrderedDict(sorted(report.items(), key=lambda t: t[0]))

    def excel_report(self, fname='ip_report.xlsx'):

        def set_headers(hrs, sheet, r=1):
            header_font = Font(color=colors.BLUE, bold=True)
            for h in hrs:
                cell = sheet.cell(row=r, column=hrs.index(h) + 1, value=h)
                cell.font = header_font
            return sheet

        wb = Workbook()
        ws = wb.active
        ws.title = 'Report'
        headers = ['Ip Address', 'Active Ports', 'Os Detected', 'Port Scanned', 'Connection refused']
        ws = set_headers(headers, ws)
        row = 2

        for key, value in self.report.items():
            if key == 0:
                continue
            ws.cell(row=row, column=headers.index('Ip Address')+1, value=value.get("Ip"))
            ws.cell(row=row, column=headers.index('Active Ports')+1, value='-'.join(value.get("Active Ports")))
            ws.cell(row=row, column=headers.index('Os Detected')+1, value=value.get("Os Detected"))
            ws.cell(row=row, column=headers.index('Port Scanned') + 1, value='-'.join(value.get("Port Scanned")))
            ws.cell(row=row, column=headers.index('Connection refused')+1, value='-'.join(value.get("Connection Refused")))
            row += 1
        wb.save(fname)




