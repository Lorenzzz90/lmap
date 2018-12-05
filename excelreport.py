from openpyxl import Workbook
from openpyxl.styles import Font, colors
from collections import OrderedDict
import os
from datetime import datetime



class ReportCreator():
    """This class is responsible for creating the reports of the scan as an excel file and as a
    graph(yet to be implemented)"""

    def __init__(self, report, folder):
        self.report = OrderedDict(sorted(report.items(), key=lambda t: t[0]))
        self.folder = folder


    def excel_report(self):
        """Create and save the excel report"""

        def set_headers(hrs, sheet, r=2):
            header_font = Font(color=colors.BLUE, bold=True)
            for h in hrs:
                cell = sheet.cell(row=r, column=hrs.index(h) + 1, value=h)
                cell.font = header_font
            return sheet

        wb = Workbook()
        ws = wb.active
        ws.title = 'Report'
        cellp = ws.cell(row=1, column=1, value="Ports Scanned")
        portfont = Font(color=colors.RED, bold=True)
        cellp.font = portfont
        col = 2
        for key, value in self.report.items():
            if key == 1:
                pass
            cell = ws.cell(row=1, column=col, value='-'.join(value.get('Port Scanned')))
            font = Font(color=colors.RED, bold=True)
            cell.font = font
        headers = ['Ip Address', 'Active Ports', 'Banners', 'Os Detected', 'Connection refused']
        ws = set_headers(headers, ws)
        row = 3

        for key, value in self.report.items():
            ws.cell(row=row, column=headers.index('Ip Address') + 1, value=value.get("Ip"))
            ws.cell(row=row, column=headers.index('Active Ports') + 1, value='-'.join(value.get("Active Ports")))
            ws.cell(row=row, column=headers.index('Banners') + 1, value='-'.join(value.get("Banners")))
            ws.cell(row=row, column=headers.index('Os Detected') + 1, value=value.get("Os Detected"))
            ws.cell(row=row, column=headers.index('Connection refused') + 1,
                    value='-'.join(value.get("Connection Refused")))
            row += 1

        for col in ws.columns:
            max_length = 0
            column = col[0].column
            for cell in col:
                if cell.coordinate == "B1":
                    continue
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[column].width = adjusted_width

        wb.save((os.path.join(self.folder, datetime.now().strftime("%d-%m-%Y_%H:%M:%S.xlsx"))))

